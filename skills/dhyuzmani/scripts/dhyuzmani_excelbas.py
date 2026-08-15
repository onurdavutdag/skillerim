# -*- coding: utf-8 -*-
"""
dhyuzmani — Excel üreticisi.

Kadro + mesafe + yatak verisine web taramasından gelen uzman/ameliyathane
sayılarını ekleyip tek sayfalık filtrelenebilir Excel üretir; ikinci sayfada
yöntem notları durur.

Kullanım:
    python dhyuzmani_excelbas.py --json yataklar.json --tarama tarama.json \
        --cikti "output/xlsx/DHY Beyin Cerrahisi YYYYAAGG SSDD.xlsx" [--referans HATAY]

`--tarama` iki şeyi kabul eder:
  * tek JSON dosyası — doğrudan okunur
  * **klasör** — içindeki `parti_*.json` dosyaları (tarama ajanlarının kendi yazdığı partiler)
    birleştirilir, birleşik liste `output/json/dhyuzmani_tarama YYYYAAGG SSDD.json` olarak yazılır

Kayıt biçimi dhyuzmani tarama ajanlarının yazdığı hâldir:
    [{"birim": ..., "bc_uzman": int|null, "bc_kaynak": url, "ameliyathane": int|null,
      "am_kaynak": url, "not": ...}, ...]
"""
import os
import sys
import json
import glob
import difflib
import argparse
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# Türkçe büyük harf: ASCII upper() 'i'yi 'I' yapar, "istanbul" vurgusu tutmaz
TR_BUYUK = str.maketrans("iıüöçşğ", "İIÜÖÇŞĞ")


def buyuk(s):
    return s.translate(TR_BUYUK).upper()


BASLIKLAR = ["#", "Birim", "İl", "Karayolu (km)", "Yaklaşık mı", "Yatak",
             "Uzman", "Ameliyathane", "Dönemler", "Kadro", "Genel Kura", "Not"]
GENISLIK = {1: 5, 2: 58, 3: 16, 4: 13, 5: 10, 6: 8, 7: 8, 8: 13, 9: 26, 10: 7, 11: 11, 12: 50}


SAYI_ALANLARI = {"bc_uzman": "bc_kaynak", "ameliyathane": "am_kaynak"}


def kayit_birlestir(a, b):
    """Aynı birim iki partide çıktıysa: dolu değer null'a yeğlenir, çakışma nota yazılır."""
    sonuc = dict(a)
    notlar = [n for n in (a.get("not"), b.get("not")) if n]
    for alan, kaynak in SAYI_ALANLARI.items():
        av, bv = a.get(alan), b.get(alan)
        if av is None and bv is not None:
            sonuc[alan] = bv
            sonuc[kaynak] = b.get(kaynak)
        elif av is not None and bv is not None and av != bv:
            notlar.append(f"çakışma: {alan} {av} ≠ {bv} (ilki alındı)")
    sonuc["not"] = " | ".join(dict.fromkeys(notlar)) or None
    return sonuc


def parti_birlestir(klasor):
    """Klasördeki parti_*.json dosyalarını tek listede toplar."""
    yollar = sorted(glob.glob(os.path.join(klasor, "parti_*.json")))
    if not yollar:
        raise SystemExit(f"HATA - {klasor} altında parti_*.json bulunamadı")
    toplam = {}
    for yol in yollar:
        with open(yol, encoding="utf-8") as f:
            veri = json.load(f)
        for kayit in veri:
            ad = kayit.get("birim")
            if not ad:
                continue
            toplam[ad] = kayit_birlestir(toplam[ad], kayit) if ad in toplam else dict(kayit)
    return list(toplam.values()), yollar


def birlesik_yol(cikti):
    """Birleşik tarama JSON'unun varsayılan yeri: <output>/json/dhyuzmani_tarama YYYYAAGG SSDD.json"""
    ad = f"dhyuzmani_tarama {datetime.now().strftime('%Y%m%d %H%M')}.json"
    kls = os.path.dirname(os.path.abspath(cikti))
    if os.path.basename(kls).lower() == "xlsx":
        return os.path.join(os.path.dirname(kls), "json", ad)
    return os.path.join(kls, ad)


def tarama_esle(birimler, tarama):
    """Tarama kayıtlarını birim adlarına bağlar (birebir, sonra bulanık)."""
    kalan = {t["birim"]: t for t in tarama}
    eslesme = {}
    for b in birimler:
        if b in kalan:
            eslesme[b] = kalan.pop(b)
            continue
        aday = difflib.get_close_matches(b, list(kalan), n=1, cutoff=0.75)
        if aday:
            eslesme[b] = kalan.pop(aday[0])
    return eslesme, list(kalan)


def excel_yaz(kayitlar, cikti, referans, tarama_dosyasi=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DHY Kadrolari"
    ws.append(BASLIKLAR)

    baslik_dolgu = PatternFill("solid", fgColor="1F4E79")
    for hucre in ws[1]:
        hucre.font = Font(bold=True, color="FFFFFF")
        hucre.fill = baslik_dolgu
        hucre.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, k in enumerate(kayitlar, 1):
        ws.append([
            i, k["birim"], k.get("il", ""), k.get("km"),
            "~" if k.get("yaklasik") else "",
            k.get("yatak") if k.get("yatak") is not None else "?",
            k.get("bc_uzman") if k.get("bc_uzman") is not None else "?",
            k.get("ameliyathane") if k.get("ameliyathane") is not None else "?",
            ", ".join(k.get("donemler", [])), k.get("kadro", ""),
            k.get("genel_kura") or "", k.get("not") or "",
        ])

    ince = Side(style="thin", color="D9D9D9")
    kenar = Border(left=ince, right=ince, top=ince, bottom=ince)
    vurgu = PatternFill("solid", fgColor="E2EFDA")
    for satir in ws.iter_rows(min_row=2):
        for hucre in satir:
            hucre.border = kenar
        if satir[2].value == buyuk(referans):
            for hucre in satir:
                hucre.fill = vurgu

    for sutun, genislik in GENISLIK.items():
        ws.column_dimensions[get_column_letter(sutun)].width = genislik
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"

    ns = wb.create_sheet("Notlar")
    for n in [
        f"Referans şehir: {referans} (karayolu mesafesi bu şehre göredir).",
        "Mesafe: KGM İller Arası Mesafe Cetveli; '~' işaretli satırlar ilçe türetmesidir (±%10).",
        "Yatak: KHGM 2.-3. Basamak Tesis Listesi (02.02.2023) tescilli yatak sütunu.",
        "UYARI: KHGM listesi 6 Şubat 2023 depreminden öncedir; Hatay ve Kahramanmaraş'ta fiilî kapasite sapabilir.",
        "Uzman ve Ameliyathane: bu çalıştırmadaki taze web taraması (hastane resmî siteleri, MHRS agregatörleri, haberler).",
        "Uzman sayısı atama/ayrılmalarla sık değişir; '?' = doğrulanabilir kaynak bulunamadı.",
        "Ameliyathane hastane genelinin toplam salon sayısıdır, branşa özel değildir.",
        f"Tarama ham verisi (kaynak URL'lerle): {tarama_dosyasi or '-'}",
    ]:
        ns.append([n])
    ns.column_dimensions["A"].width = 120

    os.makedirs(os.path.dirname(os.path.abspath(cikti)), exist_ok=True)
    wb.save(cikti)
    return ws.max_row - 1


def main():
    ap = argparse.ArgumentParser(description="dhyuzmani Excel çıktısı")
    ap.add_argument("--json", required=True, help="dhyuzmani_yatakesle.py çıktısı")
    ap.add_argument("--tarama", help="Web taraması JSON'u ya da parti_*.json klasörü")
    ap.add_argument("--birlesik", help="Klasör modunda birleşik JSON'un yazılacağı yol")
    ap.add_argument("--cikti", required=True)
    ap.add_argument("--referans", default="HATAY")
    args = ap.parse_args()

    with open(args.json, encoding="utf-8") as f:
        kayitlar = json.load(f)

    tarama_dosyasi = args.tarama
    if args.tarama:
        if os.path.isdir(args.tarama):
            tarama, yollar = parti_birlestir(args.tarama)
            tarama_dosyasi = args.birlesik or birlesik_yol(args.cikti)
            os.makedirs(os.path.dirname(os.path.abspath(tarama_dosyasi)), exist_ok=True)
            with open(tarama_dosyasi, "w", encoding="utf-8") as f:
                json.dump(tarama, f, ensure_ascii=False, indent=1)
            print(f"parti birleştirildi: {len(yollar)} dosya -> {len(tarama)} kayıt")
            print(f"birleşik tarama: {tarama_dosyasi}")
        else:
            with open(args.tarama, encoding="utf-8") as f:
                tarama = json.load(f)
        eslesme, artan = tarama_esle([k["birim"] for k in kayitlar], tarama)
        for k in kayitlar:
            t = eslesme.get(k["birim"])
            if t:
                k["bc_uzman"] = t.get("bc_uzman")
                k["ameliyathane"] = t.get("ameliyathane")
                k["not"] = t.get("not")
        if artan:
            print("UYARI - tabloya bağlanamayan tarama kaydı:", artan)
        print(f"tarama eşleşen: {len(eslesme)}/{len(kayitlar)}")

    satir = excel_yaz(kayitlar, args.cikti, args.referans, tarama_dosyasi)
    u = sum(1 for k in kayitlar if k.get("bc_uzman") is not None)
    a = sum(1 for k in kayitlar if k.get("ameliyathane") is not None)
    y = sum(1 for k in kayitlar if k.get("yatak") is not None)
    print(f"Excel yazıldı: {args.cikti}")
    print(f"satır: {satir} | yatak dolu: {y} | uzman dolu: {u} | ameliyathane dolu: {a}")


if __name__ == "__main__":
    main()
