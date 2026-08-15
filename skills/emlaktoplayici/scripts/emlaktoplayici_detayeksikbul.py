# -*- coding: utf-8 -*-
"""Detayi HENUZ toplanmamis ilan numaralarini bulur (ikinci gecisin girdisi).

Detay taramasi ilan basina 8-30 sn suren (siteye gore degisir), blok yiyebilen
uzun bir istir (bkz. references/emlaktoplayici-r-tarayici-teknigi.md §2-3). Blok gelince tarama
yarida kalir ve ikinci gecise "hangi ilanlar kaldi?" sorusuyla baslanir. Bu soru
tarama ajaninin kendi localStorage defterine BIRAKILMAZ - ajan dusmus olabilir,
sekme yenilenmis olabilir, iki ayri gecis birikmis olabilir. Cevap her zaman
elimizdeki iki gercek kaynaktan hesaplanir: Excel'deki ilan numaralari ve o ana
kadar diske yazilmis detay JSON('lar)i.

Bir ilan "tamam" sayilir:
  - detay JSON'unda anahtari varsa (ajan o sayfayi gezmis demektir), VEYA
  - Excel satirinda Aciklama ya da Tapu Durumu hucresi zaten doluysa
    (onceki bir detayekle gecisi islemis demektir).
`--kati` ile daha sert olculur: JSON kaydi olsa bile alanlarinin hepsi bossa
ilan yeniden taranir.

Girdi : mevcut .xlsx  +  0..n detay JSON (sozluk ya da {"ilanlar": [...]} bicimi)
Cikti : stdout'a JSON dizi (eksik ilan numaralari, Excel'deki sirayla)
        stderr'e sayisal ozet; istege bagli --cikti ile dosyaya da yazilir

Kullanim:
    python -B emlaktoplayici_detayeksikbul.py --xlsx "output/xlsx/... .xlsx"
    python -B emlaktoplayici_detayeksikbul.py --xlsx "... .xlsx" --detay "output/json/... .json"
"""
import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from emlaktoplayici_detayekle import sozluge_cevir  # noqa: E402  (ayni klasor, tek dogru)

# Excel'de "bu hucre bos" sayilan degerler - detayekle.py ile ayni sozlesme.
BOS_SAYILAN = (None, "", "Belirtilmemis", "Belirtilmemiş")
# Detay kaydinin dolulugu bu alanlara bakilarak olculur (--kati).
DETAY_ALANLARI = ("tapu_durumu", "aciklama", "bina_yasi", "bina_yasi_bant", "toplam_kat", "isitma")


def _anahtar(deger):
    """Basligi karsilastirilabilir hâle indirger (exceloku.anahtar ile ayni mantik).

    Turkce tuzagi: "İlan No".lower() -> "i" + U+0307 (birlesen nokta); nokta
    atilmadan "ILAN NO" / "İlan No" / "Ilan No" birbirini bulamaz.
    """
    if deger is None:
        return ""
    return str(deger).strip().lower().replace("̇", "")


def sutun_indeksi(ws, *etiketler):
    istenen = {_anahtar(e) for e in etiketler}
    for c in range(1, ws.max_column + 1):
        if _anahtar(ws.cell(1, c).value) in istenen:
            return c
    return None


def sayfayi_bul(wb, istenen):
    if istenen:
        if istenen not in wb.sheetnames:
            sys.exit("Sayfa yok: {} (mevcut: {})".format(istenen, ", ".join(wb.sheetnames)))
        return wb[istenen]
    for ad in wb.sheetnames:
        if ad not in ("Ozet", "Kunye", "Özet", "Künye"):
            return wb[ad]
    sys.exit("Ana sayfa bulunamadi.")


def detaylari_oku(yollar):
    """Birden cok detay JSON'unu tek sozlukte birlestirir. Sonraki dosya oncekini ezer."""
    toplu = {}
    for yol in yollar:
        p = Path(yol)
        if not p.exists():
            sys.exit("Detay JSON yok: {}".format(p))
        try:
            toplu.update(sozluge_cevir(json.loads(p.read_text(encoding="utf-8"))))
        except json.JSONDecodeError as e:
            sys.exit("Detay JSON bozuk: {} ({})".format(p, e))
    return toplu


def doluluk(kayit):
    if not isinstance(kayit, dict):
        return 0
    return sum(1 for alan in DETAY_ALANLARI if kayit.get(alan) not in (None, ""))


def main():
    ap = argparse.ArgumentParser(description="Detayi eksik ilan numaralarini basar.")
    ap.add_argument("--xlsx", required=True, help="Ilan numaralarinin okunacagi .xlsx")
    ap.add_argument("--detay", nargs="*", default=[], help="0..n detay JSON dosyasi")
    ap.add_argument("--sayfa", default=None, help="Ana sayfanin adi (varsayilan: ilk veri sayfasi)")
    ap.add_argument("--kati", action="store_true",
                    help="JSON kaydi bos olan ilani da eksik say (varsayilan: anahtar varsa tamam)")
    ap.add_argument("--cikti", default=None, help="Eksik numaralari bu dosyaya da yaz (JSON dizi)")
    a = ap.parse_args()

    xlsx = Path(a.xlsx)
    if not xlsx.exists():
        sys.exit("Excel yok: {}".format(xlsx))
    try:
        wb = load_workbook(xlsx, read_only=True)
    except PermissionError:
        sys.exit("Dosya Excel'de ACIK olabilir, okunamadi: {}".format(xlsx))
    ws = sayfayi_bul(wb, a.sayfa)

    ilan_no_sutunu = sutun_indeksi(ws, "Ilan No", "İlan No")
    if not ilan_no_sutunu:
        sys.exit("'Ilan No' sutunu bulunamadi - bu dosya emlaktoplayici ciktisi mi?")
    aciklama_sutunu = sutun_indeksi(ws, "Aciklama", "Açıklama")
    tapu_sutunu = sutun_indeksi(ws, "Tapu Durumu")

    detay = detaylari_oku(a.detay)

    eksik, excelden_tamam, jsondan_tamam, bos_kayit = [], 0, 0, 0
    toplam = 0

    for satir in ws.iter_rows(min_row=2, values_only=True):
        ham = satir[ilan_no_sutunu - 1]
        if ham is None:
            continue
        toplam += 1
        ilan_no = str(ham).strip()

        excelde_var = False
        for s in (aciklama_sutunu, tapu_sutunu):
            if s and satir[s - 1] not in BOS_SAYILAN:
                excelde_var = True
                break

        kayit = detay.get(ilan_no)
        if kayit is not None and a.kati and doluluk(kayit) == 0:
            bos_kayit += 1
            kayit = None

        if kayit is not None:
            jsondan_tamam += 1
        elif excelde_var:
            excelden_tamam += 1
        else:
            eksik.append(ilan_no)

    cikti = json.dumps(eksik, ensure_ascii=False)
    print(cikti)
    if a.cikti:
        hedef = Path(a.cikti)
        hedef.parent.mkdir(parents=True, exist_ok=True)
        hedef.write_text(cikti, encoding="utf-8")

    print("TOPLAM: {} | EKSIK: {} | TAMAM: {} (JSON {} + Excel {})".format(
        toplam, len(eksik), jsondan_tamam + excelden_tamam, jsondan_tamam, excelden_tamam),
        file=sys.stderr)
    if a.kati and bos_kayit:
        print("KATI KIP: bos kaydi olan {} ilan yeniden taranacak".format(bos_kayit), file=sys.stderr)
    if a.detay and not detay:
        print("UYARI: verilen detay JSON'lari bos - tum ilanlar eksik sayildi", file=sys.stderr)
    if a.cikti:
        print("YAZILDI: {}".format(a.cikti), file=sys.stderr)


if __name__ == "__main__":
    main()
