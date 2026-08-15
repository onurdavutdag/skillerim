# -*- coding: utf-8 -*-
"""Oz-denetim: varlik butunlugu + tum script zinciri, sentetik veriyle.

Kullanim:
    python -B emlaktoplayici_dogrula.py            # ag YOK (fay indirilmez)
    python -B emlaktoplayici_dogrula.py --ag       # fay indirmesini de dener

Cikis kodu 0 = hepsi gecti. Ag gerektiren tek test --ag ile acilir; varsayilan
kosu tamamen cevrimdisidir ki CI'da ve ucaktayken de kosabilsin.
"""
import argparse
import json
import re
import subprocess
import sys
import tempfile
from pathlib import Path

BURASI = Path(__file__).resolve().parent
VARLIK = BURASI.parent / "assets"
sys.path.insert(0, str(BURASI))

import emlaktoplayici_depremskorla as skorla_mod  # noqa: E402
import emlaktoplayici_excelbas as excel_mod       # noqa: E402

GECTI, KALDI = [], []


def kontrol(ad, kosul, ayrinti=""):
    (GECTI if kosul else KALDI).append(ad)
    print("  {} {}{}".format("OK  " if kosul else "HATA", ad, (" -> " + ayrinti) if ayrinti and not kosul else ""))


def kos(script, *args):
    r = subprocess.run([sys.executable, "-B", str(BURASI / script)] + [str(a) for a in args],
                       capture_output=True, text=True, encoding="utf-8", errors="replace")
    return r.returncode, (r.stdout or "") + (r.stderr or "")


def ilan(no, site, ilce, fiyat, m2, oda, **ek):
    k = {"ilan_no": no, "site": site, "baslik": "test {}".format(no), "fiyat": fiyat,
         "m2_brut": m2, "m2_net": None, "oda": oda, "il": "Hatay", "ilce": ilce,
         "semt": "", "tarih": "2026-08-10", "link": "https://t/{}".format(no),
         "bina_yasi": None, "kat": None, "toplam_kat": None, "isitma": None,
         "tapu_durumu": None, "aciklama": None}
    k.update(ek)
    return k


def paket(site, ilanlar, **meta):
    m = {"site": site, "kategori": "satilik-daire", "konum": "hatay", "arama_url": "http://t",
         "tarama_zamani": "2026-08-14T19:00:00", "sitedeki_ilan_sayisi": len(ilanlar),
         "taranan_sayfa": "1/1", "atlanan": 0, "blok_yedi_mi": False, "notlar": "sentetik"}
    m.update(meta)
    return {"meta": m, "ilanlar": ilanlar}


def yaz(yol, veri):
    Path(yol).write_text(json.dumps(veri, ensure_ascii=False), encoding="utf-8")


# ------------------------------------------------------------------------ testler

def on_madde(metin, alan):
    """Frontmatter'dan tek bir alani ceker (tirnaklari soyar)."""
    m = re.search(r"^{}\s*:\s*(.+?)\s*$".format(alan), metin, re.MULTILINE)
    return m.group(1).strip().strip('"\'') if m else None


def adlandirmayi_dogrula():
    """N kurallari - klasoreditplugin/references/adlandirma-kurali.md.

    Resmi plugin-ad-denetle.py'nin --skill kipi bu skill'i de denetler; buradaki
    kopya, oz-denetimin TEK komutla ve klasoredit kurulu olmayan makinede de
    kosabilmesi icin durur. Iki denetim celisirse resmi script kazanir.
    """
    print("\n[0] Adlandirma (N kurallari)")
    kok = BURASI.parent
    skill_adi = kok.name

    metin = (kok / "SKILL.md").read_text(encoding="utf-8")
    ad = on_madde(metin, "name")
    kontrol("N3: SKILL.md name = klasor adi", ad == skill_adi, "{} != {}".format(ad, skill_adi))
    kontrol("N9: skill adi saf ASCII kucuk harf",
            bool(re.fullmatch(r"[a-z0-9]+(-[a-z0-9]+)*", skill_adi or "")), skill_adi)
    kontrol("N4: skill adi ayracsiz tek kelime", "-" not in skill_adi, skill_adi)

    aciklama = on_madde(metin, "description") or ""
    kontrol("N11: skill description <= 1024", len(aciklama) <= 1024, "{} karakter".format(len(aciklama)))

    ajanlar = sorted((kok / "agents").glob("*.md"))
    kontrol("agents/ klasorunde ajan var", bool(ajanlar))
    for yol in ajanlar:
        dosya_adi = yol.stem
        icerik = yol.read_text(encoding="utf-8")
        a_ad = on_madde(icerik, "name")
        a_aciklama = on_madde(icerik, "description") or ""
        kontrol("N5: {} name = dosya adi".format(dosya_adi), a_ad == dosya_adi,
                "{} != {}".format(a_ad, dosya_adi))
        kontrol("N6: {} oneki <skill>-s-".format(dosya_adi),
                dosya_adi.startswith(skill_adi + "-s-"), dosya_adi)
        kontrol("N6: {} sahipligi bildiriyor (skills:)".format(dosya_adi),
                skill_adi in (on_madde(icerik, "skills") or ""), on_madde(icerik, "skills"))
        kontrol("N7: {} rolu bos degil".format(dosya_adi),
                len(dosya_adi.split("-s-", 1)[-1]) > 0)
        kontrol("N9: {} saf ASCII kucuk harf".format(dosya_adi),
                bool(re.fullmatch(r"[a-z0-9]+(-[a-z0-9]+)*", dosya_adi)))
        kontrol("N10: {} uzunlugu 3-50".format(dosya_adi), 3 <= len(dosya_adi) <= 50,
                str(len(dosya_adi)))
        kontrol("N11: {} description <= 1024".format(dosya_adi), len(a_aciklama) <= 1024,
                "{} karakter".format(len(a_aciklama)))

    for yol in sorted(BURASI.glob("*.py")):
        kontrol("N12: {} adi <skill>_<rol>.py".format(yol.name),
                bool(re.fullmatch(re.escape(skill_adi) + r"_[a-z0-9]+\.py", yol.name)), yol.name)


def varliklari_dogrula():
    print("\n[1] Varlik butunlugu")

    hasar = json.loads((VARLIK / "hatay_ilce_hasar.json").read_text(encoding="utf-8"))
    ilceler = hasar["ilceler"]
    kontrol("hasar tablosu 15 ilce", len(ilceler) == 15, str(len(ilceler)))
    for alan in ("yikik", "acil", "agir"):
        toplam = sum(d[alan] for d in ilceler.values())
        beklenen = hasar["_toplam_kontrol"][alan]
        kontrol("hasar toplami tutuyor: {}".format(alan), toplam == beklenen,
                "{} != {}".format(toplam, beklenen))
    nufus = sum(d["nufus_2022"] for d in ilceler.values())
    kontrol("nufus toplami tutuyor", nufus == hasar["_toplam_kontrol"]["nufus_2022"],
            str(nufus))
    kontrol("hasar tablosunda null yok",
            all(d.get(a) is not None for d in ilceler.values()
                for a in ("yikik", "acil", "agir", "nufus_2022")))

    koord = json.loads((VARLIK / "ilce_koordinat.json").read_text(encoding="utf-8"))["ilceler"]
    kontrol("koordinat tablosu 15 ilce", len(koord) == 15, str(len(koord)))
    kontrol("koordinatlar Hatay kutusunda",
            all(35.8 <= v["lat"] <= 37.1 and 35.7 <= v["lon"] <= 36.7 for v in koord.values()))
    kontrol("hasar ve koordinat ayni ilce kumesi", set(koord) == set(ilceler),
            str(set(koord) ^ set(ilceler)))

    lut = json.loads((VARLIK / "tdth_renk_lut.json").read_text(encoding="utf-8"))
    kontrol("LUT semasi gecerli", isinstance(lut.get("tablo"), list))
    if not lut["tablo"]:
        print("       not: LUT bos - PGA bileseni bilerek bos birakiliyor (uydurma yok)")


def deprem_dogrula():
    print("\n[2] Deprem skoru")
    hasar = skorla_mod.HasarTablosu()
    fay = skorla_mod.FayHatlari(indir=False)
    pga = skorla_mod.PgaTablosu()
    merkezler = {skorla_mod.sadelestir(k): v for k, v in
                 json.loads((VARLIK / "ilce_koordinat.json").read_text(encoding="utf-8"))["ilceler"].items()}
    t = (hasar, fay, pga, merkezler)

    yuksek = skorla_mod.skorla(ilan("1", "s", "Antakya", 3000000, 100, "2+1",
                                    bina_yasi=45, toplam_kat=8), t, 2026)
    dusuk = skorla_mod.skorla(ilan("2", "s", "Erzin", 3000000, 100, "2+1",
                                   bina_yasi=2, toplam_kat=3), t, 2026)
    kontrol("yuksek riskli ilan > dusuk riskli ilan", yuksek["skor"] > dusuk["skor"],
            "{} vs {}".format(yuksek["skor"], dusuk["skor"]))
    kontrol("skorlar 0-10 araliginda", all(0 <= x["skor"] <= 10 for x in (yuksek, dusuk)))
    kontrol("guven etiketi var", yuksek["guven"].endswith("/6 bilesen)"), yuksek["guven"])
    kontrol("PGA bileseni BOS (LUT yok)", yuksek["bilesenler"]["pga"] is None)
    # Fay: veri onbellekte VARSA hesaplanir, YOKSA bos kalir - ama asla uydurulmaz.
    # (Onceki bir --ag kosusu ~/.claude/.cache altina indirmis olabilir.)
    fay_puan = yuksek["bilesenler"]["fay"]
    if fay.parcalar:
        kontrol("fay verisi varken bilesen gecerli araliktta",
                isinstance(fay_puan, (int, float)) and 0 <= fay_puan <= skorla_mod.AZAMI["fay"],
                "{} (durum: {})".format(fay_puan, fay.durum))
    else:
        kontrol("fay verisi yokken bilesen BOS (uydurulmuyor)", fay_puan is None,
                "{} (durum: {})".format(fay_puan, fay.durum))
    kontrol("neden metni dolu", bool(yuksek["neden"]))

    yetersiz = skorla_mod.skorla(ilan("3", "s", "Bilinmeyenilce", 3000000, 100, "2+1"), t, 2026)
    kontrol("bilesen yetersizse skor URETILMEZ", yetersiz["skor"] is None
            and yetersiz["neden"] == "Yetersiz veri")

    for beyan, taban in (("orta hasarli", 8.0), ("agir hasarli", 8.0)):
        k = ilan("4", "s", "Erzin", 3000000, 100, "2+1", bina_yasi=1, toplam_kat=2,
                 aciklama="temiz daire, {} raporu var".format(beyan))
        r = skorla_mod.skorla(k, t, 2026)
        kontrol("SERT KURAL: '{}' skoru >= {}".format(beyan, taban), r["skor"] >= taban,
                str(r["skor"]))

    k = ilan("5", "s", "Erzin", 3000000, 100, "2+1", bina_yasi=40, toplam_kat=9,
             aciklama="HASARSIZ raporlu, guclendirme gerekmiyor")
    r = skorla_mod.skorla(k, t, 2026)
    kontrol("hasarsiz beyani skoru DUSURUR", r["bilesenler"]["beyan"] < 0, str(r["bilesenler"]["beyan"]))

    # --- beyan sinir durumlari (olumsuzlama + en agiri secme)
    print("  -- beyan sinir durumlari --")
    b = skorla_mod.bilesen_beyan({"baslik": "", "aciklama": "bina agir hasar almamistir, saglamdir"})
    kontrol("olumsuzlanan 'agir hasar' SERT KURALI TETIKLEMEZ",
            b[0] == 0.0 and b[2] is False, str(b))
    b = skorla_mod.bilesen_beyan({"baslik": "", "aciklama": "orta hasarli degildir"})
    kontrol("olumsuzlanan 'orta hasarli' beyan uretmez", b[0] == 0.0, str(b))
    b = skorla_mod.bilesen_beyan({"baslik": "", "aciklama": "hasarsiz raporlu daire"})
    kontrol("'hasarsiz' hasar kaliplarina YAKALANMAZ", b[0] == -1.5 and b[2] is False, str(b))
    b = skorla_mod.bilesen_beyan({"baslik": "", "aciklama": "az hasarli ama guclendirilmis bina"})
    kontrol("coklu beyanda EN AGIRI kazanir", b[0] == 0.5, str(b))
    b = skorla_mod.bilesen_beyan({"baslik": "", "aciklama": "deprem hasar kaydi yoktur"})
    kontrol("'hasar kaydi yoktur' beyan uretmez", b[0] == 0.0, str(b))
    b = skorla_mod.bilesen_beyan({"baslik": "", "aciklama": "orta hasarli, guclendirme yapilmadi"})
    kontrol("olumsuzlama CUMLECIK sinirinda kaliyor (beyan ayakta)",
            b[0] == 2.5 and b[2] is True, str(b))

    # --- yonetmelik sinir durumlari
    print("  -- yonetmelik sinir durumlari --")
    p, n = skorla_mod.bilesen_yonetmelik({"bina_yasi": 2026, "bina_yasi_bant": None}, 2026)
    kontrol("yil/yas karismasi (bina_yasi=2026) BOS + uyari",
            p is None and n and "supheli" in n, "{} / {}".format(p, n))
    p, n = skorla_mod.bilesen_yonetmelik({"bina_yasi": None, "bina_yasi_bant": "51 ve üzeri"}, 2026)
    kontrol("ucu acik bant Neden metninde 'None' YOK",
            p == 3.5 and n and "None" not in n, "{} / {}".format(p, n))
    kontrol("toplam_kat=0 kat bileseni uretmez",
            skorla_mod.bilesen_kat({"toplam_kat": 0}) == (None, None))
    kontrol("toplam_kat=-1 kat bileseni uretmez",
            skorla_mod.bilesen_kat({"toplam_kat": -1}) == (None, None))

    # TR harf duyarsizligi: "Kırıkhan" ile "KIRIKHAN" ayni ilceyi bulmali
    a = hasar.puan("Kırıkhan")[0]
    b = hasar.puan("KIRIKHAN")[0]
    kontrol("TR buyuk/kucuk harf duyarsiz ilce eslesmesi", a is not None and a == b,
            "{} vs {}".format(a, b))

    # --- bina yasi BANDI (sahibinden bandli veriyor: "11-15 arası")
    print("  -- bina yasi bandi --")
    kontrol("bant ayiklama: '11-15 arası'", skorla_mod._yas_araligi("11-15 arası") == (11, 15))
    kontrol("bant ayiklama: '31 ve üzeri'", skorla_mod._yas_araligi("31 ve üzeri") == (31, None))
    kontrol("bant ayiklama: '4'", skorla_mod._yas_araligi("4") == (4, 4))
    kontrol("bant ayiklama: sayi yoksa None", skorla_mod._yas_araligi("belirtilmemis") is None)

    # 2026'da 11-15 yas -> 2011-2015 -> hepsi 2007-2018 bandinda -> KESIN 1.0
    p, n = skorla_mod.bilesen_yonetmelik({"bina_yasi": None, "bina_yasi_bant": "11-15 arası"}, 2026)
    kontrol("tek yonetmelige oturan bant KESIN cozuluyor", p == 1.0, "{} / {}".format(p, n))
    kontrol("kesin bantta 'ust sinir' notu YOK", n and "ust sinir" not in n, str(n))

    # 2026'da 5-10 yas -> 2016-2021 -> 2007-2018 (1.0) ile >=2019 (0.5) arasinda yayiliyor
    p, n = skorla_mod.bilesen_yonetmelik({"bina_yasi": None, "bina_yasi_bant": "5-10 arası"}, 2026)
    kontrol("yayilan bantta YUKSEK riskli sinir aliniyor", p == 1.0, "{} / {}".format(p, n))
    kontrol("yayilan bantta 'ust sinir' notu VAR", n and "ust sinir" in n, str(n))

    # 2026'da 31+ yas -> <=1995 -> 1976-1997 (3.0) ile <=1975 (3.5) arasinda; ust sinir 3.5
    p, n = skorla_mod.bilesen_yonetmelik({"bina_yasi": None, "bina_yasi_bant": "31 ve üzeri"}, 2026)
    kontrol("ucu acik bantta en yuksek risk aliniyor", p == 3.5, "{} / {}".format(p, n))

    # Kesin yas her zaman banda tercih edilir
    p, _ = skorla_mod.bilesen_yonetmelik({"bina_yasi": 2, "bina_yasi_bant": "31 ve üzeri"}, 2026)
    kontrol("kesin yas bandi EZIYOR", p == 0.5, str(p))


def zincir_dogrula(gecici):
    print("\n[3] Uctan uca zincir")
    ayni = dict(ilce="Antakya", fiyat=2950000, m2=140, oda="3+1")
    yaz(gecici / "a.json", paket("sahibinden", [
        ilan("A1", "sahibinden", bina_yasi=12, toplam_kat=5, **ayni),
        ilan("A2", "sahibinden", ilce="Erzin", fiyat=2800000, m2=100, oda="2+1", bina_yasi=3, toplam_kat=4),
    ]))
    yaz(gecici / "b.json", paket("hepsiemlak", [
        ilan("B1", "hepsiemlak", ilce="Antakya", fiyat=2960000, m2=141, oda="3+1",
             bina_yasi=12, toplam_kat=5, tapu_durumu="Kat Mulkiyetli"),
    ], atlanan=2))
    yaz(gecici / "c.json", paket("emlakjet", [
        ilan("C1", "emlakjet", ilce="Antakya", fiyat=2955000, m2=140, oda="3+1"),
    ], blok_yedi_mi=True, atlanan=7))

    kod, cikti = kos("emlaktoplayici_depremskorla.py", "--kayit", gecici / "a.json",
                     "--cikti", gecici / "skorlu.json", "--fay-indirme-yok")
    kontrol("depremskorla calisti", kod == 0, cikti.strip()[-200:])

    kod, cikti = kos("emlaktoplayici_mesafehesapla.py", "--kayit", gecici / "skorlu.json",
                     "--cikti", gecici / "mesafeli.json", "--referans", "Antakya")
    kontrol("mesafehesapla calisti", kod == 0, cikti.strip()[-200:])
    if kod == 0:
        m = json.loads((gecici / "mesafeli.json").read_text(encoding="utf-8"))["ilanlar"]
        antakya = [x for x in m if x["ilce"] == "Antakya"][0]
        kontrol("Antakya'nin Antakya'ya mesafesi 0", antakya["mesafe_km"] == 0.0,
                str(antakya["mesafe_km"]))
        kontrol("yaklasik mesafe ISARETLENDI", antakya.get("mesafe_yaklasik") is True)

    xlsx = gecici / "cikti.xlsx"
    kod, cikti = kos("emlaktoplayici_excelbas.py", "--kayit", gecici / "mesafeli.json",
                     gecici / "b.json", gecici / "c.json", "--cikti", xlsx)
    kontrol("excelbas calisti", kod == 0, cikti.strip()[-300:])
    if kod != 0:
        return

    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    kontrol("uc sayfa var", len(wb.sheetnames) == 3 and {"Ozet", "Kunye"} <= set(wb.sheetnames),
            str(wb.sheetnames))
    ws = wb[wb.sheetnames[0]]
    basliklar = [c.value for c in ws[1]]
    kontrol("zorunlu basliklar yerinde",
            {"Ilan Basligi", "Fiyat (TL)", "Ilce", "Ilan No", "Kaynak", "Link"} <= set(basliklar),
            str(basliklar))
    kontrol("dondurulmus satir", ws.freeze_panes == "A2", str(ws.freeze_panes))
    kontrol("otomatik filtre", bool(ws.auto_filter.ref), str(ws.auto_filter.ref))

    # 4 ham kayit -> 3 site birbirine benziyor: A1+B1+C1 tek satir, A2 ayri = 2 satir
    kontrol("sitelerarasi tekillestirme calisti", ws.max_row - 1 == 2,
            "{} satir".format(ws.max_row - 1))

    fiyat_sutunu = basliklar.index("Fiyat (TL)") + 1
    m2_sutunu = basliklar.index("m2 (Brut)") + 1
    kontrol("fiyat SAYI tipinde",
            all(isinstance(ws.cell(r, fiyat_sutunu).value, (int, float))
                for r in range(2, ws.max_row + 1)))
    kontrol("m2 SAYI tipinde",
            all(isinstance(ws.cell(r, m2_sutunu).value, (int, float))
                for r in range(2, ws.max_row + 1)))

    kaynak_sutunu = basliklar.index("Kaynak") + 1
    kaynaklar = [ws.cell(r, kaynak_sutunu).value for r in range(2, ws.max_row + 1)]
    kontrol("birlesen satir uc siteyi de listeliyor",
            any(k and k.count(",") == 2 for k in kaynaklar), str(kaynaklar))

    ozet = "\n".join(str(c.value) for satir in wb["Ozet"].iter_rows() for c in satir if c.value)
    kontrol("Ozet blok uyarisini iceriyor", "BLOK" in ozet)
    kontrol("Ozet atlanan ilani bildiriyor", "Atlanan ilan" in ozet)
    kontrol("Ozet deprem uyari metnini iceriyor", "ON ELEME" in ozet)
    kunye = "\n".join(str(c.value) for satir in wb["Kunye"].iter_rows() for c in satir if c.value)
    kontrol("Kunye deprem uyari metnini iceriyor", "ON ELEME" in kunye)

    # --- tekillestirmede SKORLU kayit kazanir (15.08.2026 duzeltmesi)
    s1 = ilan("S1", "sahibinden", "Antakya", 2950000, 140, "3+1", tarih="2026-08-12")
    s1["deprem"] = {"skor": None, "guven": "(2/6 bilesen)", "bilesenler": {}, "neden": "Yetersiz veri"}
    s2 = ilan("S2", "hepsiemlak", "Antakya", 2950000, 140, "3+1", tarih="2026-08-01")
    s2["deprem"] = {"skor": 5.0, "guven": "(4/6 bilesen)", "bilesenler": {}, "neden": "test"}
    b_kayitlar, _, _ = excel_mod.tekillestir([s1, s2])
    kontrol("birlesmede SKORLU deprem kaydi kazaniyor (daha yeni skorsuza karsi)",
            len(b_kayitlar) == 1 and (b_kayitlar[0].get("deprem") or {}).get("skor") == 5.0,
            str([(k.get("deprem") or {}).get("skor") for k in b_kayitlar]))

    # --- olasi tekrar: kimlik tutuyor, fiyat tutmuyor -> birlesmez, raporlanir
    t1 = ilan("T1", "sahibinden", "Antakya", 2500000, 120, "2+1")
    t2 = ilan("T2", "emlakjet", "Antakya", 3000000, 120, "2+1")
    o_kayitlar, _, olasi = excel_mod.tekillestir([t1, t2])
    kontrol("olasi tekrar BIRLESTIRILMIYOR ama yakalaniyor",
            len(o_kayitlar) == 2 and len(olasi) == 1, "{} kayit, {} suphe".format(
                len(o_kayitlar), len(olasi)))
    yaz(gecici / "olasi.json", paket("sahibinden", [t1]))
    yaz(gecici / "olasi2.json", paket("emlakjet", [t2]))
    kod, cikti = kos("emlaktoplayici_excelbas.py", "--kayit", gecici / "olasi.json",
                     gecici / "olasi2.json", "--cikti", gecici / "olasi.xlsx")
    if kod == 0:
        wbo = load_workbook(gecici / "olasi.xlsx")
        ozet_o = "\n".join(str(c.value) for satir in wbo["Ozet"].iter_rows() for c in satir if c.value)
        kontrol("Ozet olasi tekrari LISTELIYOR", "Olasi tekrar" in ozet_o and "T1" in ozet_o)
    else:
        kontrol("Ozet olasi tekrari LISTELIYOR", False, cikti.strip()[-200:])

    # --- detay ekleme (kat=-1: bodrum, eksi isaretin tum turu sag cikmasi test edilir)
    yaz(gecici / "detay.json", {"A1": {"tapu_durumu": "Kat Mulkiyetli", "aciklama": "x" * 40000,
                                       "kat": -1}})
    kod, cikti = kos("emlaktoplayici_detayekle.py", "--xlsx", xlsx, "--detay", gecici / "detay.json")
    kontrol("detayekle calisti", kod == 0, cikti.strip()[-300:])
    if kod == 0:
        wb2 = load_workbook(xlsx)
        b2 = [c.value for c in wb2[wb2.sheetnames[0]][1]]
        kontrol("Aciklama sutunu EN SONDA", b2[-1] == "Aciklama", str(b2[-1]))
        kontrol("detayekle 'Bulundugu Kat' sutunu acti", "Bulundugu Kat" in b2, str(b2))
        uzun = [wb2[wb2.sheetnames[0]].cell(r, len(b2)).value
                for r in range(2, wb2[wb2.sheetnames[0]].max_row + 1)]
        kontrol("uzun aciklama kirpildi",
                any(v and "kisaltildi" in str(v) for v in uzun))
    # Ikinci kosu IDEMPOTENT olmali: sutun tekrar acilmaz, dolu hucre korunur.
    onceki_sutun_sayisi = load_workbook(xlsx)[wb.sheetnames[0]].max_column
    kod, cikti = kos("emlaktoplayici_detayekle.py", "--xlsx", xlsx, "--detay", gecici / "detay.json")
    wb3 = load_workbook(xlsx)
    kontrol("detayekle ikinci kosu IDEMPOTENT (hata vermez)", kod == 0, cikti.strip()[-200:])
    kontrol("ikinci kosu sutun COGALTMIYOR",
            wb3[wb3.sheetnames[0]].max_column == onceki_sutun_sayisi,
            "{} -> {}".format(onceki_sutun_sayisi, wb3[wb3.sheetnames[0]].max_column))
    kontrol("ikinci kosu dolu hucreyi KORUYOR", "yazilan hucre: 0" in cikti, cikti.strip()[-200:])

    # --- detayeksikbul: ikinci gecisin girdisi. Su an yalniz A1'in detayi var.
    kod, cikti = kos("emlaktoplayici_detayeksikbul.py", "--xlsx", xlsx)
    kontrol("detayeksikbul calisti", kod == 0, cikti.strip()[-200:])
    kontrol("detayeksikbul EKSIGI buluyor (Excel'deki dolu hucreden okuyarak)",
            "EKSIK: 1" in cikti and '"A2"' in cikti, cikti.strip()[-200:])

    # --- ajan liste semasini yazarsa is copre gitmemeli: detayekle onu da cevirir
    yaz(gecici / "detay_liste.json",
        {"meta": {"site": "sahibinden"},
         "ilanlar": [{"ilan_no": "A2", "tapu_durumu": "Kat Irtifakli", "isitma": "Kombi"}]})
    kod, cikti = kos("emlaktoplayici_detayekle.py", "--xlsx", xlsx,
                     "--detay", gecici / "detay_liste.json")
    kontrol("detayekle LISTE bicimini de kabul ediyor",
            kod == 0 and "yazilan hucre: 2" in cikti, cikti.strip()[-200:])

    kod, cikti = kos("emlaktoplayici_detayeksikbul.py", "--xlsx", xlsx)
    kontrol("detayeksikbul: detay tamamlaninca EKSIK 0",
            kod == 0 and "EKSIK: 0" in cikti, cikti.strip()[-200:])

    # --- exceloku: xlsx -> kayit JSON -> tekrar xlsx (tam tur)
    geri = gecici / "geri.json"
    kod, cikti = kos("emlaktoplayici_exceloku.py", "--xlsx", xlsx, "--cikti", geri)
    kontrol("exceloku calisti", kod == 0, cikti.strip()[-300:])
    if kod == 0:
        veri = json.loads(geri.read_text(encoding="utf-8"))
        satir_sayisi = load_workbook(xlsx)[wb.sheetnames[0]].max_row - 1
        kontrol("exceloku tum satirlari okudu",
                len(veri.get("ilanlar", [])) == satir_sayisi,
                "{} != {}".format(len(veri.get("ilanlar", [])), satir_sayisi))
        kontrol("exceloku zorunlu alanlari dolduruyor",
                all(all(k.get(a) not in (None, "") for a in
                        ("ilan_no", "site", "baslik", "fiyat", "ilce", "link"))
                    for k in veri.get("ilanlar", [])),
                str(veri.get("ilanlar", [{}])[0]))
        kontrol("exceloku detay alanlarini koruyor",
                any(k.get("tapu_durumu") for k in veri["ilanlar"])
                and any(k.get("aciklama") for k in veri["ilanlar"]))
        kontrol("exceloku EKSI kati koruyor (bodrum -1)",
                any(k.get("kat") == -1 for k in veri["ilanlar"]),
                str([k.get("kat") for k in veri["ilanlar"]]))
        # Turkce baslikli tablo da okunmali: "Ilce".lower() birlesen nokta uretir
        # ve naif eslestirme sutunu sessizce dusurur (ilk kosuda oldu).
        wbt = load_workbook(xlsx)
        wst = wbt[wbt.sheetnames[0]]
        for c in range(1, wst.max_column + 1):
            if wst.cell(1, c).value == "Ilce":
                wst.cell(1, c, "İlçe")
            elif wst.cell(1, c).value == "Ilan Basligi":
                wst.cell(1, c, "İlan Başlığı")
        turkce_xlsx = gecici / "turkce.xlsx"
        wbt.save(turkce_xlsx)
        kod, cikti = kos("emlaktoplayici_exceloku.py", "--xlsx", turkce_xlsx, "--cikti", gecici / "geri_tr.json")
        tr = json.loads((gecici / "geri_tr.json").read_text(encoding="utf-8")) if kod == 0 else {}
        kontrol("exceloku TURKCE baslikli tabloyu da okuyor",
                kod == 0 and all(k.get("ilce") and k.get("baslik") for k in tr.get("ilanlar", [{}])),
                cikti.strip()[-200:])

        tur_xlsx = gecici / "tur.xlsx"
        kod, cikti = kos("emlaktoplayici_excelbas.py", "--kayit", geri, "--cikti", tur_xlsx)
        kontrol("tam tur: geri okunan JSON tekrar Excel basiyor", kod == 0, cikti.strip()[-300:])
        if kod == 0:
            wbtur = load_workbook(tur_xlsx)
            btur = [c.value for c in wbtur[wbtur.sheetnames[0]][1]]
            kontrol("tam turda Bulundugu Kat / Isitma KAYBOLMUYOR",
                    "Bulundugu Kat" in btur and "Isitma" in btur, str(btur))

    # --- fark
    kod, cikti = kos("emlaktoplayici_farkcikar.py", "--onceki", gecici / "a.json",
                     "--simdiki", gecici / "a.json", "--cikti", gecici / "fark0.json")
    kontrol("ayni girdi -> 0 yeni / 0 degisen / 0 kalkan",
            kod == 0 and "0 yeni, 0 fiyati degisen, 0 kalkan" in cikti, cikti.strip()[-200:])

    degisik = json.loads((gecici / "a.json").read_text(encoding="utf-8"))
    degisik["ilanlar"][0]["fiyat"] = 2700000
    degisik["ilanlar"].append(ilan("A3", "sahibinden", "Belen", 2900000, 90, "2+1"))
    degisik["ilanlar"].pop(1)
    yaz(gecici / "a2.json", degisik)
    kod, cikti = kos("emlaktoplayici_farkcikar.py", "--onceki", gecici / "a.json",
                     "--simdiki", gecici / "a2.json", "--cikti", gecici / "fark1.json")
    kontrol("1 yeni / 1 degisen / 1 kalkan dogru sayildi",
            kod == 0 and "1 yeni, 1 fiyati degisen, 1 kalkan" in cikti, cikti.strip()[-200:])

    # --- detaybirlestir: detay -> KAYIT JSON -> skor zinciri (SKILL.md Adim 4'un yolu).
    # Aciklama ve toplam_kat skorlayiciya ancak bu yoldan ulasir - 15.08.2026'da
    # bu halka yoktu ve tablo hem aciklamasiz hem skorsuz cikti.
    print("  -- detaybirlestir zinciri --")
    yaz(gecici / "detay_skor.json",
        {"A1": {"aciklama": "guzel daire, hasarsiz raporlu", "toplam_kat": 5, "kat": 2,
                "isitma": "Kombi", "bina_yasi": 13},
         "A2": {"aciklama": "sahibinden orta hasarli, guclendirme yapilmadi"}})
    kod, cikti = kos("emlaktoplayici_detaybirlestir.py", "--kayit", gecici / "a.json",
                     "--detay", gecici / "detay_skor.json", "--cikti", gecici / "birlesik.json")
    kontrol("detaybirlestir calisti", kod == 0, cikti.strip()[-300:])
    if kod == 0:
        bir = json.loads((gecici / "birlesik.json").read_text(encoding="utf-8"))
        a1 = [k for k in bir["ilanlar"] if k["ilan_no"] == "A1"][0]
        kontrol("detaybirlestir aciklama + kat + isitma isliyor",
                a1.get("aciklama") and a1.get("kat") == 2 and a1.get("isitma") == "Kombi",
                str({x: a1.get(x) for x in ("aciklama", "kat", "isitma")}))
        kontrol("detaybirlestir DOLU alani koruyor (bina_yasi 12 != detay 13)",
                a1.get("bina_yasi") == 12, str(a1.get("bina_yasi")))
        kontrol("detaybirlestir celiskiyi RAPORLUYOR", "CELISKI" in cikti, cikti.strip()[-200:])

        kod, cikti = kos("emlaktoplayici_depremskorla.py", "--kayit", gecici / "birlesik.json",
                         "--cikti", gecici / "birlesik_skorlu.json", "--fay-indirme-yok")
        kontrol("detay sonrasi depremskorla calisti", kod == 0, cikti.strip()[-200:])
        if kod == 0:
            sk = json.loads((gecici / "birlesik_skorlu.json").read_text(encoding="utf-8"))["ilanlar"]
            a1s = [k for k in sk if k["ilan_no"] == "A1"][0]["deprem"]
            a2s = [k for k in sk if k["ilan_no"] == "A2"][0]["deprem"]
            kontrol("skorlayici detaydan gelen BEYANI goruyor (hasarsiz -1.5)",
                    a1s["bilesenler"]["beyan"] == -1.5, str(a1s["bilesenler"]))
            kontrol("skorlayici detaydan gelen TOPLAM KATI goruyor",
                    a1s["bilesenler"]["kat"] is not None, str(a1s["bilesenler"]))
            kontrol("detaydan gelen 'orta hasarli' SERT KURALI tetikliyor",
                    a2s["skor"] is not None and a2s["skor"] >= 8.0, str(a2s["skor"]))


def sema_dogrula(gecici):
    print("\n[4] Sema koruyuculari")
    eksik = paket("sahibinden", [{"ilan_no": "X", "site": "sahibinden", "baslik": "y",
                                  "fiyat": 100, "link": "http://z"}])  # 'ilce' yok
    yaz(gecici / "eksik.json", eksik)
    kod, cikti = kos("emlaktoplayici_excelbas.py", "--kayit", gecici / "eksik.json",
                     "--cikti", gecici / "olmaz.xlsx")
    kontrol("zorunlu alan eksikse REDDEDIYOR", kod != 0 and "ilce" in cikti, cikti.strip()[-160:])

    (gecici / "bozuk.json").write_text("{ bu json degil", encoding="utf-8")
    kod, cikti = kos("emlaktoplayici_excelbas.py", "--kayit", gecici / "bozuk.json",
                     "--cikti", gecici / "olmaz2.xlsx")
    kontrol("bozuk JSON'da anlamli hata veriyor", kod != 0 and "Bozuk JSON" in cikti,
            cikti.strip()[-160:])

    metin_fiyat = paket("sahibinden", [{"ilan_no": "Y", "site": "sahibinden", "baslik": "y",
                                        "fiyat": "2.950.000 TL", "ilce": "Antakya",
                                        "link": "http://z"}])
    yaz(gecici / "metinfiyat.json", metin_fiyat)
    kod, cikti = kos("emlaktoplayici_excelbas.py", "--kayit", gecici / "metinfiyat.json",
                     "--cikti", gecici / "olmaz3.xlsx")
    kontrol("metin fiyati anlamli hatayla REDDEDIYOR (TypeError degil)",
            kod != 0 and "sayi degil" in cikti and "Traceback" not in cikti, cikti.strip()[-160:])

    kod, cikti = kos("emlaktoplayici_excelbas.py", "--kayit", gecici / "eslesmez*.json",
                     "--cikti", gecici / "olmaz4.xlsx")
    kontrol("genislememis glob deseninde anlamli hata (PowerShell yolu)",
            kod != 0 and "eslesmedi" in cikti, cikti.strip()[-160:])

    # Bozuk JSON diger scriptlerde de traceback DEGIL anlamli hata vermeli.
    xlsx_var = gecici / "cikti.xlsx"  # zincir testinden kalan gercek dosya
    kod, cikti = kos("emlaktoplayici_detayekle.py", "--xlsx", xlsx_var,
                     "--detay", gecici / "bozuk.json")
    kontrol("detayekle bozuk JSON'u anlamli reddediyor",
            kod != 0 and "bozuk" in cikti.lower() and "Traceback" not in cikti, cikti.strip()[-160:])
    kod, cikti = kos("emlaktoplayici_detayeksikbul.py", "--xlsx", xlsx_var,
                     "--detay", gecici / "bozuk.json")
    kontrol("detayeksikbul bozuk JSON'u anlamli reddediyor",
            kod != 0 and "bozuk" in cikti.lower() and "Traceback" not in cikti, cikti.strip()[-160:])
    kod, cikti = kos("emlaktoplayici_farkcikar.py", "--onceki", gecici / "bozuk.json",
                     "--simdiki", gecici / "bozuk.json", "--cikti", gecici / "olmaz5.json")
    kontrol("farkcikar bozuk JSON'u anlamli reddediyor",
            kod != 0 and "Bozuk JSON" in cikti and "Traceback" not in cikti, cikti.strip()[-160:])
    yaz(gecici / "yanlis_sema.json", {"foo": 1})
    kod, cikti = kos("emlaktoplayici_farkcikar.py", "--onceki", gecici / "yanlis_sema.json",
                     "--simdiki", gecici / "yanlis_sema.json", "--cikti", gecici / "olmaz6.json")
    kontrol("farkcikar yanlis semayi sessizce 'hepsi yeni' SAYMIYOR",
            kod != 0 and "ilanlar" in cikti, cikti.strip()[-160:])
    kod, cikti = kos("emlaktoplayici_mesafehesapla.py", "--kayit", gecici / "bozuk.json",
                     "--cikti", gecici / "olmaz7.json", "--referans", "Antakya")
    kontrol("mesafehesapla bozuk JSON'u anlamli reddediyor",
            kod != 0 and "bozuk" in cikti.lower() and "Traceback" not in cikti, cikti.strip()[-160:])
    kod, cikti = kos("emlaktoplayici_depremskorla.py", "--kayit", gecici / "bozuk.json",
                     "--cikti", gecici / "olmaz8.json", "--fay-indirme-yok")
    kontrol("depremskorla bozuk ile eksik dosyayi AYIRT ediyor",
            kod != 0 and "bozuk" in cikti.lower(), cikti.strip()[-160:])


def dokuman_sayisi_dogrula():
    """README ve SKILL.md'deki denetim sayisi GERCEK sayiyla esit mi?

    Sayi iki kez bayatladi (79 -> 89 -> 101 derken gercek 102 idi). Kontrol
    buraya alindi: sayi bayatlarsa oz-denetim DUSER, bakim kurali kendini zorlar.
    Not: bu fonksiyondaki iki kontrol de toplama dahildir; --ag'in +2'si degildir
    (dokumanlar varsayilan kosunun sayisini yazar).
    """
    print("\n[5] Dokuman-sayi esitligi")
    kok = BURASI.parent
    beklenen = len(GECTI) + len(KALDI) + 2  # asagidaki iki kontrol de sayilir

    readme = (kok / "README.md").read_text(encoding="utf-8")
    m = re.search(r"\*\*(\d+)\s+denetim\*\*", readme)
    kontrol("README denetim sayisi guncel ({})".format(beklenen),
            bool(m) and int(m.group(1)) == beklenen,
            "{} != {}".format(m.group(1) if m else "bulunamadi", beklenen))

    skill = (kok / "SKILL.md").read_text(encoding="utf-8")
    m = re.search(r"(\d+)\s+kontrol", skill)
    kontrol("SKILL.md kontrol sayisi guncel ({})".format(beklenen),
            bool(m) and int(m.group(1)) == beklenen,
            "{} != {}".format(m.group(1) if m else "bulunamadi", beklenen))


def ag_dogrula():
    print("\n[6] Ag testi (--ag)")
    fay = skorla_mod.FayHatlari(indir=True)
    kontrol("GEM diri fay verisi indi ve yuklendi", bool(fay.parcalar), fay.durum)
    if fay.parcalar:
        # Antakya Dogu Anadolu Fay Zonu'na cok yakin; Erzin belirgin daha uzak olmali
        pa = fay.puan(36.2023, 36.1613)
        pe = fay.puan(36.9500, 36.2000)
        kontrol("Antakya faya Erzin'den yakin", pa[0] >= pe[0], "{} vs {}".format(pa, pe))


def main():
    ap = argparse.ArgumentParser(description="emlaktoplayici oz-denetimi")
    ap.add_argument("--ag", action="store_true", help="Ag gerektiren testleri de kos")
    a = ap.parse_args()

    print("emlaktoplayici oz-denetim")
    adlandirmayi_dogrula()
    varliklari_dogrula()
    deprem_dogrula()
    with tempfile.TemporaryDirectory() as td:
        gecici = Path(td)
        zincir_dogrula(gecici)
        sema_dogrula(gecici)
    dokuman_sayisi_dogrula()
    if a.ag:
        ag_dogrula()

    print("\n{} gecti, {} kaldi".format(len(GECTI), len(KALDI)))
    if KALDI:
        print("KALAN:")
        for ad in KALDI:
            print("  -", ad)
        sys.exit(1)
    print("hepsi gecti.")


if __name__ == "__main__":
    main()
