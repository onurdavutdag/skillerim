# -*- coding: utf-8 -*-
"""Mevcut bir Excel ciktisini geri okuyup kayit JSON'una cevirir (excelbas'in tersi).

Neden gerekli: paketteki her sey kayit JSON'u uzerinden konusur - tekillestirme,
deprem skoru, mesafe, fark raporu, hepsi o semayi okur. Elde yalniz Excel varsa
(kayit JSON'u kaybolmus ya da tablo tek seferlik bir script'le uretilmis olabilir)
o veri zincire ancak JSON'a donerse girer. Tipik kullanim: eski bir taramayi yeni
site taramalariyla BIRLESTIRIP tabloyu yeniden basmak.

Turkce ve ASCII basliklarin ikisini de tanir - tablo hangi script'le uretilmis
olursa olsun okunur.

Girdi : .xlsx  (Ana sayfa + varsa Kunye sayfasi)
Cikti : {"meta": {...}, "ilanlar": [...]}  -> excelbas'in bekledigi sema

Kullanim:
    python -B emlaktoplayici_exceloku.py --xlsx "output/xlsx/... .xlsx" --cikti kayit.json
    python -B emlaktoplayici_exceloku.py --xlsx "... .xlsx" --cikti k.json --site sahibinden
"""
import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

# Baslik -> sema alani. Anahtarlar kucuk harfe indirgenmis hâlleriyle aranir;
# hem Turkce (tek seferlik script'lerin urettigi) hem ASCII (excelbas) taninir.
BASLIK_ALAN = {
    "ilan basligi": "baslik", "ilan başlığı": "baslik",
    "fiyat (tl)": "fiyat",
    "m2 (brut)": "m2_brut", "m² (brüt)": "m2_brut",
    "m2 (net)": "m2_net", "m² (net)": "m2_net",
    "oda": "oda",
    "tapu durumu": "tapu_durumu",
    "ilce": "ilce", "ilçe": "ilce",
    "semt": "semt",
    "bina yasi": "bina_yasi", "bina yaşı": "bina_yasi",
    "bina yasi bandi": "bina_yasi_bant", "bina yaşı bandı": "bina_yasi_bant",
    "toplam kat": "toplam_kat",
    "kat": "kat",
    "isitma": "isitma", "isıtma": "isitma",
    "mesafe (km)": "mesafe_km",
    "ilan tarihi": "tarih", "i̇lan tarihi": "tarih",
    "ilan no": "ilan_no", "i̇lan no": "ilan_no",
    "link": "link",
    "aciklama": "aciklama", "açıklama": "aciklama",
}
# excelbas'in turettigi sutunlar - geri okunmaz, kayittan yeniden hesaplanir.
TURETILMIS = {"tl/m2", "tl/m²"}
# Deprem skoru sutunlari: skor JSON'da ic ice bir sozluktur, duz sutunlardan
# guvenilir sekilde geri kurulamaz -> atlanir, gerekirse depremskorla yeniden kosar.
# "Kat" basligi bu blokta da var; blok varsa "Kat" oraya aittir (asagida ayiklanir).
DEPREM_BASLIKLARI = {"deprem risk", "guven", "güven", "yonetmelik", "yönetmelik",
                     "ilce hasar", "ilçe hasar", "fay", "pga", "beyan", "neden"}
BOS_SAYILAN = (None, "", "Belirtilmemis", "Belirtilmemiş", "?")

KUNYE_META = {
    "arama url": "arama_url",
    "veri çekim tarihi": "tarama_zamani", "veri cekim tarihi": "tarama_zamani",
    "tarama zamani": "tarama_zamani", "tarama zamanı": "tarama_zamani",
    "sitedeki ilan sayısı": "sitedeki_ilan_sayisi", "sitedeki ilan sayisi": "sitedeki_ilan_sayisi",
    "taranan sayfa": "taranan_sayfa",
    "filtre": "filtre_metni",
    "kaynak": "kaynak_metni",
}


def anahtar(deger):
    """Basligi karsilastirilabilir hâle indirger.

    Turkce tuzagi: "Ilce".lower() -> "i" + U+0307 (birlesen nokta), yani
    "İlçe".lower() != "ilçe". Birlesen noktayi atmadan hicbir Turkce baslik
    eslesmez - bu script ilk kosusunda tam olarak bundan iki sutun kaybetti.
    """
    if deger is None:
        return ""
    return str(deger).strip().lower().replace("̇", "")


def sayi(deger):
    """'2.950.000 TL' / 2950000 / '140' -> int; cozulemezse None."""
    if deger in BOS_SAYILAN:
        return None
    if isinstance(deger, (int, float)):
        return int(deger)
    temiz = re.sub(r"[^\d]", "", str(deger))
    return int(temiz) if temiz else None


def tarihe_cevir(deger):
    if deger in BOS_SAYILAN:
        return None
    if isinstance(deger, (datetime, date)):
        return deger.strftime("%Y-%m-%d")
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", str(deger).strip())
    if m:
        return m.group(0)
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", str(deger).strip())
    if m:
        return "{}-{}-{}".format(m.group(3), m.group(2), m.group(1))
    return str(deger).strip() or None


def sayfayi_bul(wb, istenen):
    if istenen:
        if istenen not in wb.sheetnames:
            sys.exit("Sayfa yok: {} (mevcut: {})".format(istenen, ", ".join(wb.sheetnames)))
        return wb[istenen]
    for ad in wb.sheetnames:
        if ad not in ("Ozet", "Kunye", "Özet", "Künye"):
            return wb[ad]
    sys.exit("Ana sayfa bulunamadi.")


def kunye_oku(wb):
    """Kunye sayfasindaki etiket/deger ciftlerinden meta kurar."""
    meta = {}
    ad = "Kunye" if "Kunye" in wb.sheetnames else ("Künye" if "Künye" in wb.sheetnames else None)
    if not ad:
        return meta
    for satir in wb[ad].iter_rows(values_only=True):
        if not satir or satir[0] is None or len(satir) < 2:
            continue
        alan = KUNYE_META.get(anahtar(satir[0]))
        if alan and satir[1] not in BOS_SAYILAN:
            meta[alan] = satir[1]
    fiyatlar = re.findall(r"[\d.]{7,}", str(meta.get("filtre_metni", "")))
    if len(fiyatlar) >= 2:
        meta["filtre"] = {"fiyat_min": sayi(fiyatlar[0]), "fiyat_max": sayi(fiyatlar[1])}
    return meta


def main():
    ap = argparse.ArgumentParser(description="Excel ciktisini kayit JSON'una cevirir.")
    ap.add_argument("--xlsx", required=True, help="Okunacak .xlsx")
    ap.add_argument("--cikti", required=True, help="Yazilacak kayit JSON")
    ap.add_argument("--site", default=None,
                    help="'Kaynak' sutunu yoksa kullanilacak site adi (ornek: sahibinden)")
    ap.add_argument("--sayfa", default=None, help="Ana sayfanin adi (varsayilan: ilk veri sayfasi)")
    ap.add_argument("--il", default=None, help="'il' alani icin deger (varsayilan: Kunye'den, yoksa bos)")
    a = ap.parse_args()

    xlsx = Path(a.xlsx)
    if not xlsx.exists():
        sys.exit("Excel yok: {}".format(xlsx))
    try:
        wb = load_workbook(xlsx, read_only=True)
    except PermissionError:
        sys.exit("Dosya Excel'de ACIK olabilir, okunamadi: {}".format(xlsx))
    ws = sayfayi_bul(wb, a.sayfa)

    basliklar = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    kucuk = [anahtar(b) for b in basliklar]
    deprem_var = "deprem risk" in kucuk

    yerlesim, atlanan_sutun = {}, []
    for i, k in enumerate(kucuk):
        if not k or k in TURETILMIS:
            continue
        if k in DEPREM_BASLIKLARI or (deprem_var and k == "kat"):
            atlanan_sutun.append(basliklar[i])
            continue
        if k == "kaynak":
            yerlesim[i] = "_kaynak"
        elif k in BASLIK_ALAN:
            yerlesim[i] = BASLIK_ALAN[k]
        else:
            atlanan_sutun.append(basliklar[i])

    meta = kunye_oku(wb)
    il = a.il or ("Hatay" if "hatay" in str(meta.get("kaynak_metni", "")).lower() else None)

    ilanlar, sitesiz = [], 0
    for satir in ws.iter_rows(min_row=2, values_only=True):
        if not any(h not in BOS_SAYILAN for h in satir):
            continue
        k = {}
        for i, alan in yerlesim.items():
            if i >= len(satir):
                continue
            deger = satir[i]
            if deger in BOS_SAYILAN:
                continue
            if alan in ("fiyat", "m2_brut", "m2_net", "bina_yasi", "toplam_kat", "kat"):
                deger = sayi(deger)
            elif alan == "tarih":
                deger = tarihe_cevir(deger)
            elif alan in ("ilan_no", "oda", "ilce", "semt", "baslik", "link", "tapu_durumu",
                          "isitma", "bina_yasi_bant", "aciklama", "_kaynak"):
                deger = str(deger).strip()
            k[alan] = deger

        kaynak = k.pop("_kaynak", None)
        site = (kaynak.split(",")[0].strip() if kaynak else None) or a.site
        if not site:
            sitesiz += 1
        k["site"] = site
        if il:
            k.setdefault("il", il)
        # Sema sozlesmesi: bulunamayan alan silinmez, null kalir (0 != null).
        for alan in ("m2_net", "bina_yasi", "bina_yasi_bant", "toplam_kat", "kat",
                     "isitma", "tapu_durumu", "aciklama"):
            k.setdefault(alan, None)
        ilanlar.append(k)

    if sitesiz:
        sys.exit("{} satirda site belirlenemedi ('Kaynak' sutunu yok). --site ile ver.".format(sitesiz))

    meta.setdefault("site", ilanlar[0]["site"] if ilanlar else (a.site or ""))
    meta.setdefault("sitedeki_ilan_sayisi", len(ilanlar))
    meta.setdefault("atlanan", 0)
    meta.setdefault("blok_yedi_mi", False)
    meta["kaynak_dosya"] = xlsx.name
    meta["notlar"] = "Excel'den geri okundu ({} satir).".format(len(ilanlar))

    Path(a.cikti).write_text(
        json.dumps({"meta": meta, "ilanlar": ilanlar}, ensure_ascii=False, indent=1),
        encoding="utf-8")

    dolu = lambda alan: sum(1 for k in ilanlar if k.get(alan) not in (None, ""))
    print("okunan ilan: {} | yazildi: {}".format(len(ilanlar), a.cikti))
    print("dolu alanlar: tapu_durumu {} | aciklama {} | bina_yasi {} | bina_yasi_bant {}".format(
        dolu("tapu_durumu"), dolu("aciklama"), dolu("bina_yasi"), dolu("bina_yasi_bant")))
    if atlanan_sutun:
        print("atlanan sutun (turetilmis/deprem): {}".format(", ".join(str(s) for s in atlanan_sutun)))
        if deprem_var:
            print("NOT: deprem skoru geri okunmaz - gerekiyorsa depremskorla.py yeniden kosulur.")


if __name__ == "__main__":
    main()
