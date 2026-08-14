# -*- coding: utf-8 -*-
"""Tarama JSON'larini tek bir filtrelenebilir Excel dosyasina basar.

Girdi : bir veya daha cok tarama JSON'u (references/excel-sozlesmesi.md §1 semasi)
Cikti : uc sayfali .xlsx  ->  Ana | Ozet | Kunye

Kullanim:
    python -B emlaktoplayici_excelbas.py --kayit sahibinden.json hepsiemlak.json \
        --cikti "output/xlsx/hatay satilik daire 20260814 2130.xlsx"

Bu script hicbir veriyi GOMULU tutmaz ve hicbir dosya yolunu SABITLEMEZ; ikisi de
turedigi tek seferlik script'in kusuruydu.
"""
import argparse
import json
import statistics
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASLIK_DOLGU = PatternFill("solid", fgColor="1F3864")
BASLIK_YAZI = Font(color="FFFFFF", bold=True)
LINK_YAZI = Font(color="0563C1", underline="single")
HUCRE_SINIRI = 32000  # Excel siniri 32.767; guvenli pay birakilir

ZORUNLU_ALANLAR = ("ilan_no", "site", "baslik", "fiyat", "ilce", "link")

UYARI_METNI = (
    "Deprem risk skoru bir ON ELEME aracidir, muhendislik degerlendirmesi degildir. "
    "Kamuya acik ilan verisi ve bolgesel verilerden hesaplanir; binanin kendisi incelenmemistir. "
    "Satin alma oncesi sunlar sarttir: adresle hasar tespit sorgusu, tapuda 'riskli yapi' serhi kontrolu, "
    "ruhsat ve iskan yilinin belgeyle dogrulanmasi, yapi denetim dosyasinin incelenmesi ve gerekirse "
    "bir insaat muhendisinden yapisal degerlendirme alinmasi. "
    "Skorun 0 olmasi binanin guvenli oldugu anlamina GELMEZ."
)


# ----------------------------------------------------------------- okuma / dogrulama

def kayit_oku(yollar):
    """Tarama JSON'larini okur, semayi dogrular, (metalar, ilanlar) dondurur."""
    metalar, ilanlar = [], []
    for yol in yollar:
        p = Path(yol)
        if not p.exists():
            sys.exit("Dosya yok: {}".format(p))
        try:
            veri = json.loads(p.read_text(encoding="utf-8"))
        except json.JSONDecodeError as e:
            sys.exit("Bozuk JSON ({}): {}".format(p.name, e))

        meta = veri.get("meta") or {}
        meta["_dosya"] = p.name
        metalar.append(meta)

        satirlar = veri.get("ilanlar")
        if not isinstance(satirlar, list):
            sys.exit("{}: 'ilanlar' listesi yok.".format(p.name))

        for i, k in enumerate(satirlar):
            eksik = [a for a in ZORUNLU_ALANLAR if k.get(a) in (None, "")]
            if eksik:
                sys.exit("{} kayit #{}: zorunlu alan bos -> {}".format(p.name, i, ", ".join(eksik)))
            ilanlar.append(k)

    return metalar, ilanlar


def m2_al(k):
    """Brut varsa brut, yoksa net. Ikisi de yoksa None."""
    return k.get("m2_brut") or k.get("m2_net")


def tarih_al(k):
    t = k.get("tarih")
    if not t:
        return None
    try:
        return datetime.strptime(str(t)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


# ------------------------------------------------------------------- tekillestirme

def _yakin(a, b, oran):
    if not a or not b:
        return False
    return abs(a - b) <= max(a, b) * oran


def tekillestir(ilanlar):
    """Ayni daire birden cok sitede olabilir. excel-sozlesmesi.md §5 kurali.

    Doner: (birlesik_kayitlar, birlesen_sayisi, olasi_tekrarlar)
    """
    kumeler = []          # her biri: {"kayitlar": [...], "siteler": set()}
    olasi_tekrar = []

    for k in ilanlar:
        yerlesti = False
        for kume in kumeler:
            ref = kume["kayitlar"][0]
            ayni_ilce = (k.get("ilce") or "") == (ref.get("ilce") or "")
            ayni_oda = (k.get("oda") or "") == (ref.get("oda") or "")
            m2_uyar = _yakin(m2_al(k), m2_al(ref), 0.03)
            fiyat_uyar = _yakin(k.get("fiyat"), ref.get("fiyat"), 0.02)

            tutan = sum([ayni_ilce, ayni_oda, m2_uyar, fiyat_uyar])
            if tutan == 4 and k.get("site") != ref.get("site"):
                kume["kayitlar"].append(k)
                kume["siteler"].add(k.get("site"))
                yerlesti = True
                break
            # Uc kriter tutup biri tutmuyorsa BIRLESTIRME - supheyi rapor et.
            if tutan == 3 and k.get("site") != ref.get("site"):
                olasi_tekrar.append((ref.get("ilan_no"), k.get("ilan_no")))

        if not yerlesti:
            kumeler.append({"kayitlar": [k], "siteler": {k.get("site")}})

    birlesik, birlesen = [], 0
    for kume in kumeler:
        kayitlar = kume["kayitlar"]
        if len(kayitlar) > 1:
            birlesen += len(kayitlar) - 1
        # Dolu olan kazanir; ikisi de doluysa daha YENI tarih kazanir.
        sirali = sorted(kayitlar, key=lambda x: (tarih_al(x) or date.min), reverse=True)
        ana = dict(sirali[0])
        for digeri in sirali[1:]:
            for alan, deger in digeri.items():
                if ana.get(alan) in (None, "") and deger not in (None, ""):
                    ana[alan] = deger
        ana["_kaynaklar"] = ", ".join(sorted(s for s in kume["siteler"] if s))
        birlesik.append(ana)

    return birlesik, birlesen, olasi_tekrar


# ------------------------------------------------------------------------- sutunlar

def sutunlari_kur(kayitlar):
    """Veride ne varsa o sutun eklenir. Bos sutun uretilmez."""
    var = lambda alan: any(k.get(alan) not in (None, "") for k in kayitlar)
    deprem_var = any(isinstance(k.get("deprem"), dict) for k in kayitlar)

    sutunlar = [
        ("Ilan Basligi", "baslik", 52),
        ("Fiyat (TL)", "fiyat", 14),
        ("m2 (Brut)", "_m2", 10),
        ("TL/m2", "_tl_m2", 12),
        ("Oda", "oda", 8),
    ]
    if var("tapu_durumu"):
        sutunlar.append(("Tapu Durumu", "tapu_durumu", 17))
    sutunlar += [
        ("Ilce", "ilce", 13),
        ("Semt", "semt", 14),
    ]
    if var("bina_yasi"):
        sutunlar.append(("Bina Yasi", "bina_yasi", 10))
    if var("bina_yasi_bant"):
        sutunlar.append(("Bina Yasi Bandi", "bina_yasi_bant", 15))
    if var("toplam_kat"):
        sutunlar.append(("Toplam Kat", "toplam_kat", 11))
    if deprem_var:
        sutunlar += [
            ("Deprem Risk", "_d_skor", 12),
            ("Guven", "_d_guven", 10),
            ("Yonetmelik", "_d_yonetmelik", 11),
            ("Ilce Hasar", "_d_hasar", 11),
            ("Fay", "_d_fay", 8),
            ("PGA", "_d_pga", 8),
            ("Kat", "_d_kat", 8),
            ("Beyan", "_d_beyan", 9),
            ("Neden", "_d_neden", 46),
        ]
    if var("mesafe_km"):
        sutunlar.append(("Mesafe (km)", "mesafe_km", 12))
    sutunlar += [
        ("Ilan Tarihi", "_tarih", 13),
        ("Ilan No", "ilan_no", 14),
        ("Kaynak", "_kaynaklar", 22),
        ("Link", "link", 46),
    ]
    if var("aciklama"):
        sutunlar.append(("Aciklama", "aciklama", 60))  # daima en sonda
    return sutunlar


def turetilmis_alanlar(k):
    """Hesaplanan alanlari kayda ekler."""
    m2 = m2_al(k)
    k["_m2"] = m2
    k["_tl_m2"] = round(k["fiyat"] / m2) if m2 else None
    k["_tarih"] = tarih_al(k)
    k.setdefault("_kaynaklar", k.get("site"))

    d = k.get("deprem")
    if isinstance(d, dict):
        b = d.get("bilesenler") or {}
        k["_d_skor"] = d.get("skor")
        k["_d_guven"] = d.get("guven")            # "(4/6)"
        k["_d_yonetmelik"] = b.get("yonetmelik")
        k["_d_hasar"] = b.get("ilce_hasar")
        k["_d_fay"] = b.get("fay")
        k["_d_pga"] = b.get("pga")
        k["_d_kat"] = b.get("kat")
        k["_d_beyan"] = b.get("beyan")
        k["_d_neden"] = d.get("neden")
    return k


# ---------------------------------------------------------------------- yazim

def basligi_bicimle(hucre, etiket):
    hucre.value = etiket
    hucre.fill = BASLIK_DOLGU
    hucre.font = BASLIK_YAZI
    hucre.alignment = Alignment(horizontal="center", vertical="center")


def ana_sayfa(wb, kayitlar, sutunlar, sayfa_adi):
    ws = wb.active
    ws.title = sayfa_adi[:31]  # Excel sayfa adi siniri

    for no, (etiket, _, genislik) in enumerate(sutunlar, start=1):
        basligi_bicimle(ws.cell(row=1, column=no), etiket)
        ws.column_dimensions[get_column_letter(no)].width = genislik
    ws.row_dimensions[1].height = 26

    kirpilan = 0
    for satir, k in enumerate(kayitlar, start=2):
        for no, (_, alan, _) in enumerate(sutunlar, start=1):
            deger = k.get(alan)
            if alan == "aciklama" and isinstance(deger, str) and len(deger) > HUCRE_SINIRI:
                deger = deger[:HUCRE_SINIRI] + " [...kisaltildi]"
                kirpilan += 1
            hucre = ws.cell(row=satir, column=no, value=deger)

            if alan in ("fiyat", "_tl_m2"):
                hucre.number_format = "#,##0"
            elif alan in ("_m2", "bina_yasi", "toplam_kat"):
                hucre.number_format = "0"
            elif alan in ("_d_skor", "mesafe_km") or alan.startswith("_d_") and alan != "_d_neden":
                if isinstance(deger, (int, float)):
                    hucre.number_format = "0.0"
            elif alan == "_tarih":
                hucre.number_format = "DD.MM.YYYY"
                hucre.alignment = Alignment(horizontal="center")
            elif alan == "ilan_no":
                hucre.number_format = "@"  # metin: bastaki sifir korunur
            elif alan == "link" and deger:
                hucre.hyperlink = deger
                hucre.font = LINK_YAZI
            elif alan == "aciklama":
                hucre.alignment = Alignment(vertical="top", wrap_text=False)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(get_column_letter(len(sutunlar)), len(kayitlar) + 1)
    return kirpilan


def _sayilar(kayitlar, alan):
    return [k[alan] for k in kayitlar if isinstance(k.get(alan), (int, float))]


def ozet_sayfa(wb, kayitlar, metalar, birlesen, olasi_tekrar):
    ws = wb.create_sheet("Ozet")
    yaz = lambda *h: ws.append(list(h))

    def bolum(baslik):
        ws.append([])
        r = ws.max_row + 1
        ws.cell(r, 1, baslik).font = Font(bold=True, size=12)

    # --- ilce kirilimi
    bolum("Ilce kirilimi")
    yaz("Ilce", "Adet", "Ort. Fiyat", "Ort. TL/m2", "En ucuz", "En pahali")
    for h in ws[ws.max_row]:
        h.font = Font(bold=True)

    ilceler = {}
    for k in kayitlar:
        ilceler.setdefault(k.get("ilce") or "?", []).append(k)
    for ilce in sorted(ilceler, key=lambda x: -len(ilceler[x])):
        grup = ilceler[ilce]
        fiyatlar = _sayilar(grup, "fiyat")
        tlm2 = _sayilar(grup, "_tl_m2")
        yaz(
            ilce,
            len(grup),
            round(statistics.mean(fiyatlar)) if fiyatlar else None,
            round(statistics.mean(tlm2)) if tlm2 else None,
            min(fiyatlar) if fiyatlar else None,
            max(fiyatlar) if fiyatlar else None,
        )

    # --- dagilim
    bolum("Dagilim")
    for etiket, alan in (("Fiyat (TL)", "fiyat"), ("TL/m2", "_tl_m2")):
        v = sorted(_sayilar(kayitlar, alan))
        if not v:
            continue
        ceyrek = statistics.quantiles(v, n=4) if len(v) >= 4 else [None, statistics.median(v), None]
        yaz(etiket, "min", v[0], "Q1", ceyrek[0], "medyan", ceyrek[1], "Q3", ceyrek[2], "maks", v[-1])

    # --- kaynak
    bolum("Kaynak siteler")
    site_sayisi = {}
    for k in kayitlar:
        for s in (k.get("_kaynaklar") or "").split(","):
            s = s.strip()
            if s:
                site_sayisi[s] = site_sayisi.get(s, 0) + 1
    for s, n in sorted(site_sayisi.items(), key=lambda x: -x[1]):
        yaz(s, n)
    yaz("Tekillestirmede birlesen satir", birlesen)

    # --- kapsam uyarilari
    bolum("Kapsam uyarilari")
    atlanan_toplam = sum(int(m.get("atlanan") or 0) for m in metalar)
    yaz("Atlanan ilan (toplam)", atlanan_toplam)
    for m in metalar:
        if m.get("blok_yedi_mi"):
            yaz("BLOK", "{} taramasi blok yedi - veri EKSIK".format(m.get("site")))
    if olasi_tekrar:
        yaz("Olasi tekrar (birlestirilmedi)", len(olasi_tekrar))
        for a, b in olasi_tekrar[:50]:
            yaz("", "{} <-> {}".format(a, b))
    if not atlanan_toplam and not olasi_tekrar and not any(m.get("blok_yedi_mi") for m in metalar):
        yaz("", "Uyari yok - tarama tam.")

    # --- deprem
    if any(isinstance(k.get("deprem"), dict) for k in kayitlar):
        bolum("Deprem risk skoru dagilimi")
        bantlar = [(0, 2.5, "Dusuk"), (2.5, 4.5, "Orta-dusuk"), (4.5, 6.5, "Orta"),
                   (6.5, 8.0, "Yuksek"), (8.0, 10.1, "Cok yuksek")]
        skorlar = _sayilar(kayitlar, "_d_skor")
        for alt, ust, ad in bantlar:
            yaz(ad, "{:.1f} - {:.1f}".format(alt, min(ust, 10.0)),
                sum(1 for s in skorlar if alt <= s < ust))
        yaz("Skor uretilemeyen", len(kayitlar) - len(skorlar))
        ws.append([])
        u = ws.cell(ws.max_row + 1, 1, "UYARI: " + UYARI_METNI)
        u.font = Font(bold=True, color="C00000")
        u.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=u.row, start_column=1, end_row=u.row, end_column=6)
        ws.row_dimensions[u.row].height = 90

    ws.column_dimensions["A"].width = 34
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16


def kunye_sayfa(wb, kayitlar, metalar, kirpilan):
    ws = wb.create_sheet("Kunye")
    yaz = lambda a, b: ws.append([a, b])

    for m in metalar:
        ws.append([])
        ws.cell(ws.max_row + 1, 1, "Kaynak: {}".format(m.get("site") or m.get("_dosya"))).font = Font(bold=True, size=12)
        for etiket, anahtar in (
            ("Kategori", "kategori"), ("Konum", "konum"), ("Arama URL", "arama_url"),
            ("Tarama zamani", "tarama_zamani"), ("Sitedeki ilan sayisi", "sitedeki_ilan_sayisi"),
            ("Taranan sayfa", "taranan_sayfa"), ("Atlanan ilan", "atlanan"),
            ("Blok yedi mi", "blok_yedi_mi"), ("Notlar", "notlar"),
        ):
            if m.get(anahtar) is not None:
                yaz(etiket, m[anahtar])
        f = m.get("filtre")
        if f:
            yaz("Filtre", json.dumps(f, ensure_ascii=False))

    ws.append([])
    ws.cell(ws.max_row + 1, 1, "Birlesik tablo").font = Font(bold=True, size=12)
    fiyatlar = _sayilar(kayitlar, "fiyat")
    tlm2 = _sayilar(kayitlar, "_tl_m2")
    yaz("Tabloya giren ilan", len(kayitlar))
    yaz("Ortalama fiyat", round(statistics.mean(fiyatlar)) if fiyatlar else None)
    yaz("Ortalama TL/m2", round(statistics.mean(tlm2)) if tlm2 else None)
    yaz("Aciklamasi kirpilan satir", kirpilan)
    yaz("Uretim zamani", datetime.now().strftime("%d.%m.%Y %H:%M") + " (yerel)")
    yaz("Not", "m2 degerleri ilan sahibinin beyanidir; brut/net ayrimi ilan detayinda kontrol edilmelidir.")
    if any(isinstance(k.get("deprem"), dict) for k in kayitlar):
        yaz("UYARI", UYARI_METNI)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    for h in ws["A"]:
        if h.value and not str(h.value).startswith("Kaynak:"):
            h.font = Font(bold=True)
    for h in ws["B"]:
        h.alignment = Alignment(wrap_text=True, vertical="top")


# ------------------------------------------------------------------------- akis

def main():
    ap = argparse.ArgumentParser(description="Tarama JSON'larini Excel'e basar.")
    ap.add_argument("--kayit", nargs="+", required=True, help="Bir veya cok tarama JSON'u")
    ap.add_argument("--cikti", required=True, help="Hedef .xlsx yolu")
    ap.add_argument("--sayfa-adi", default="Ilanlar", help="Ana sayfanin adi")
    ap.add_argument("--tekillestirme-yok", action="store_true",
                    help="Sitelerarasi tekillestirmeyi kapatir (ham hali gorulsun diye)")
    a = ap.parse_args()

    metalar, ham = kayit_oku(a.kayit)
    if not ham:
        sys.exit("Hic ilan yok - Excel uretilmedi.")

    if a.tekillestirme_yok:
        kayitlar, birlesen, olasi = [dict(k, _kaynaklar=k.get("site")) for k in ham], 0, []
    else:
        kayitlar, birlesen, olasi = tekillestir(ham)

    kayitlar = [turetilmis_alanlar(k) for k in kayitlar]
    kayitlar.sort(key=lambda k: (k.get("fiyat") or 0))

    sutunlar = sutunlari_kur(kayitlar)
    wb = Workbook()
    kirpilan = ana_sayfa(wb, kayitlar, sutunlar, a.sayfa_adi)
    ozet_sayfa(wb, kayitlar, metalar, birlesen, olasi)
    kunye_sayfa(wb, kayitlar, metalar, kirpilan)

    hedef = Path(a.cikti)
    hedef.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(hedef)
    except PermissionError:
        sys.exit("Dosya Excel'de ACIK olabilir, yazilamadi: {}".format(hedef))

    print("{} ham -> {} satir ({} birlesti) | {} sutun -> {}".format(
        len(ham), len(kayitlar), birlesen, len(sutunlar), hedef))
    if olasi:
        print("UYARI: {} olasi tekrar birlestirilmedi, Ozet sayfasinda listeli.".format(len(olasi)))
    for m in metalar:
        if m.get("blok_yedi_mi"):
            print("UYARI: {} blok yedi, {} ilan atlandi.".format(m.get("site"), m.get("atlanan")))


if __name__ == "__main__":
    main()
