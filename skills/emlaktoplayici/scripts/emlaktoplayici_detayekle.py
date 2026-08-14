# -*- coding: utf-8 -*-
"""Detay sayfasindan toplanan alanlari MEVCUT bir Excel dosyasina yerinde ekler.

Detay alanlari (Aciklama, Tapu Durumu, Bina Yasi, Kat) liste sayfasinda YOKTUR;
ilan sayfasina tek tek girmeyi gerektirir ve pahalidir (ilan basina >=12-15 sn,
bkz. references/tarayici-teknigi.md §5). Bu yuzden ayri gecis, ayri script.

Girdi : mevcut .xlsx  +  detay JSON  { "<ilan no>": {"tapu_durumu": ..., "aciklama": ...} }
Cikti : ayni .xlsx, yerinde guncellenir (dosya adi DEGISMEZ - guncelleme kurali)

Kullanim:
    python -B emlaktoplayici_detayekle.py --xlsx "output/xlsx/... .xlsx" --detay detay.json
"""
import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASLIK_DOLGU = PatternFill("solid", fgColor="1F3864")
BASLIK_YAZI = Font(color="FFFFFF", bold=True)
HUCRE_SINIRI = 32000

# (JSON alani, Excel basligi, genislik, nereye) - "son" = tablonun sonuna
EKLENECEK = [
    ("tapu_durumu", "Tapu Durumu", 17, "Oda"),
    ("bina_yasi", "Bina Yasi", 10, "Semt"),
    ("toplam_kat", "Toplam Kat", 11, "Semt"),
    ("aciklama", "Aciklama", 60, "son"),
]


def sutun_indeksi(ws, etiket):
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == etiket:
            return c
    return None


def basligi_bicimle(hucre, etiket):
    hucre.value = etiket
    hucre.fill = BASLIK_DOLGU
    hucre.font = BASLIK_YAZI
    hucre.alignment = Alignment(horizontal="center", vertical="center")


def sayfayi_bul(wb, istenen):
    if istenen:
        if istenen not in wb.sheetnames:
            sys.exit("Sayfa yok: {} (mevcut: {})".format(istenen, ", ".join(wb.sheetnames)))
        return wb[istenen]
    for ad in wb.sheetnames:
        if ad not in ("Ozet", "Kunye", "Özet", "Künye"):
            return wb[ad]
    sys.exit("Ana sayfa bulunamadi.")


def main():
    ap = argparse.ArgumentParser(description="Detay alanlarini mevcut Excel'e ekler.")
    ap.add_argument("--xlsx", required=True, help="Guncellenecek .xlsx")
    ap.add_argument("--detay", required=True, help="Detay JSON (anahtar = ilan no)")
    ap.add_argument("--sayfa", default=None, help="Ana sayfanin adi (varsayilan: ilk veri sayfasi)")
    ap.add_argument("--uzerine-yaz", action="store_true",
                    help="Dolu hucrelerin uzerine de yazar (varsayilan: dolu hucre KORUNUR)")
    a = ap.parse_args()
    uzerine_yaz = a.uzerine_yaz

    xlsx = Path(a.xlsx)
    if not xlsx.exists():
        sys.exit("Excel yok: {}".format(xlsx))
    detay = json.loads(Path(a.detay).read_text(encoding="utf-8"))
    if not isinstance(detay, dict):
        sys.exit("Detay JSON bir sozluk olmali: {'<ilan no>': {...}}")

    try:
        wb = load_workbook(xlsx)
    except PermissionError:
        sys.exit("Dosya Excel'de ACIK olabilir, okunamadi: {}".format(xlsx))
    ws = sayfayi_bul(wb, a.sayfa)

    ilan_no_sutunu = sutun_indeksi(ws, "Ilan No") or sutun_indeksi(ws, "İlan No")
    if not ilan_no_sutunu:
        sys.exit("'Ilan No' sutunu bulunamadi - bu dosya emlaktoplayici ciktisi mi?")

    # Hangi alanlar gercekten geliyor? Bos sutun acilmaz.
    gelen = {alan for kayit in detay.values() for alan, d in kayit.items() if d not in (None, "")}
    eklenecek = [e for e in EKLENECEK if e[0] in gelen]
    if not eklenecek:
        sys.exit("Detay JSON'da dolu alan yok - yapacak bir sey yok.")

    # Sutun zaten olabilir: ilk tarama bazi alanlari listede yakalamis olabilir
    # (or. hepsiemlak bina yasini liste sayfasinda verir). Bu durumda YENI SUTUN
    # ACILMAZ, mevcut sutunun BOS hucreleri doldurulur. Script boylece idempotenttir:
    # ikinci kosu veriyi bozmaz, "0 yeni deger" der ve gecer.
    yerlesim, mevcut = {}, []
    for alan, etiket, genislik, nereye in eklenecek:
        varsa = sutun_indeksi(ws, etiket)
        if varsa:
            mevcut.append(etiket)
        else:
            if nereye == "son":
                hedef = ws.max_column + 1
            else:
                komsu = sutun_indeksi(ws, nereye)
                hedef = (komsu + 1) if komsu else (ws.max_column + 1)
                ws.insert_cols(hedef)
            basligi_bicimle(ws.cell(1, hedef), etiket)
            ws.column_dimensions[ws.cell(1, hedef).column_letter].width = genislik
        yerlesim[alan] = etiket  # indeks kayabilir, etiketle tekrar bulunur

    ilan_no_sutunu = sutun_indeksi(ws, "Ilan No") or sutun_indeksi(ws, "İlan No")
    indeksler = {alan: sutun_indeksi(ws, etiket) for alan, etiket in yerlesim.items()}

    eslesmeyen = kirpilan = yazilan = korunan = uzerine = 0
    bos = {alan: 0 for alan in yerlesim}
    dagilim = {}
    BOS_SAYILAN = (None, "", "Belirtilmemis", "Belirtilmemiş")

    for r in range(2, ws.max_row + 1):
        ham = ws.cell(r, ilan_no_sutunu).value
        if ham is None:
            continue
        ilan_no = str(ham).strip()
        kayit = detay.get(ilan_no)

        if kayit is None:
            eslesmeyen += 1
            sutun = indeksler.get("tapu_durumu")
            if sutun and ws.cell(r, sutun).value in BOS_SAYILAN:
                ws.cell(r, sutun, "Belirtilmemis")
            continue

        for alan, sutun in indeksler.items():
            deger = kayit.get(alan)
            if deger in (None, ""):
                bos[alan] += 1
                if alan == "tapu_durumu" and ws.cell(r, sutun).value in BOS_SAYILAN:
                    ws.cell(r, sutun, "Belirtilmemis")
                continue

            # Dolu bir hucrenin uzerine, acikca istenmedikce YAZILMAZ.
            eldeki = ws.cell(r, sutun).value
            if eldeki not in BOS_SAYILAN:
                if not uzerine_yaz:
                    korunan += 1
                    continue
                uzerine += 1
            yazilan += 1

            if alan == "aciklama":
                deger = str(deger).strip()
                if len(deger) > HUCRE_SINIRI:
                    deger = deger[:HUCRE_SINIRI] + " [...kisaltildi]"
                    kirpilan += 1
                h = ws.cell(r, sutun, deger)
                h.alignment = Alignment(vertical="top", wrap_text=False)
            elif alan in ("bina_yasi", "toplam_kat"):
                try:
                    h = ws.cell(r, sutun, int(deger))
                    h.number_format = "0"
                except (TypeError, ValueError):
                    ws.cell(r, sutun, str(deger))
            else:
                v = str(deger).strip()
                ws.cell(r, sutun, v)
                if alan == "tapu_durumu":
                    dagilim[v] = dagilim.get(v, 0) + 1

    son_sutun = ws.cell(1, ws.max_column).column_letter
    ws.auto_filter.ref = "A1:{}{}".format(son_sutun, ws.max_row)
    ws.freeze_panes = "A2"

    # Kunye guncelle - hangi geciste ne toplandigi kayda gecsin
    kunye = wb["Kunye"] if "Kunye" in wb.sheetnames else (wb["Künye"] if "Künye" in wb.sheetnames else None)
    if kunye is not None:
        kunye.append([])
        kunye.append(["Detay taramasi", datetime.now().strftime("%d.%m.%Y %H:%M") + " (yerel)"])
        kunye.append(["Detayi okunan ilan", len(detay)])
        kunye.append(["Detayi alinamayan satir", eslesmeyen])
        kunye.append(["Yazilan hucre", yazilan])
        if korunan:
            kunye.append(["Korunan dolu hucre (uzerine yazilmadi)", korunan])
        if uzerine:
            kunye.append(["Uzerine yazilan hucre", uzerine])
        for alan in yerlesim:
            kunye.append(["Bos: {}".format(yerlesim[alan]), bos[alan]])
        if kirpilan:
            kunye.append(["Aciklamasi kirpilan satir", kirpilan])
        for h in kunye["A"]:
            if h.value:
                h.font = Font(bold=True)

    try:
        wb.save(xlsx)
    except PermissionError:
        sys.exit("Dosya Excel'de ACIK - kaydedilemedi. Kapatip tekrar calistir: {}".format(xlsx))

    yeni_sutun = [e for e in yerlesim.values() if e not in mevcut]
    print("yeni sutun: {} | mevcut sutun dolduruldu: {}".format(
        ", ".join(yeni_sutun) or "yok", ", ".join(mevcut) or "yok"))
    print("detay kayit: {} | yazilan hucre: {} | korunan dolu hucre: {} | uzerine yazilan: {} | "
          "eslesmeyen satir: {} | kirpilan: {}".format(
              len(detay), yazilan, korunan, uzerine, eslesmeyen, kirpilan))
    if dagilim:
        print("tapu dagilimi:", sorted(dagilim.items(), key=lambda x: -x[1]))
    print("guncellendi:", xlsx)


if __name__ == "__main__":
    main()
